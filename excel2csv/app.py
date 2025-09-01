import csv
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import load_workbook  # pip install openpyxl


def pick_excel_file() -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.update()
    path = filedialog.askopenfilename(
        title="Selecione um arquivo Excel (.xlsx)",
        filetypes=[("Excel .xlsx", "*.xlsx"), ("Todos os arquivos", "*.*")]
    )
    root.destroy()
    return Path(path) if path else None


def ask_delimiter() -> str:
    root = tk.Tk()
    root.withdraw()
    delim = simpledialog.askstring(
        "Delimitador",
        "Informe o delimitador do CSV (padrão ; )",
        initialvalue=";"
    )
    root.destroy()
    return delim or ";"


def ask_sheet_choice(sheets: list[str]) -> str | None:
    if not sheets:
        return None
    if len(sheets) == 1:
        return sheets[0]

    root = tk.Tk()
    root.withdraw()
    msg = (
        "Abas encontradas:\n- " + "\n- ".join(sheets) +
        "\n\nDigite o nome exato da aba para exportar somente ela,\n"
        "ou digite ALL para exportar TODAS."
    )
    choice = simpledialog.askstring("Escolher aba(s)", msg, initialvalue="ALL")
    root.destroy()
    return choice


def safe_cell(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", " ").replace("\n", " ").strip()


def export_sheet_to_csv(xlsx_path: Path, sheet_name: str, delimiter: str, encoding: str = "utf-8-sig") -> Path:
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada no arquivo.")
    ws = wb[sheet_name]

    out_path = xlsx_path.with_name(f"{xlsx_path.stem}__{sheet_name}.csv")
    with open(out_path, "w", encoding=encoding, newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([safe_cell(v) for v in row])
    return out_path


def main():
    xlsx_path = pick_excel_file()
    if not xlsx_path:
        print("Nenhum arquivo selecionado.")
        return

    try:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível abrir o Excel: {e}")
        return

    choice = ask_sheet_choice(sheets)
    if not choice:
        print("Nenhuma aba selecionada.")
        return

    delim = ask_delimiter()
    exported = []

    try:
        if choice.upper() == "ALL":
            for s in sheets:
                out = export_sheet_to_csv(xlsx_path, s, delim)
                exported.append(out)
        else:
            out = export_sheet_to_csv(xlsx_path, choice, delim)
            exported.append(out)

        msg = "Arquivos exportados:\n" + "\n".join(str(p.name) for p in exported)
        print(msg)
        messagebox.showinfo("Concluído", msg)
    except Exception as e:
        messagebox.showerror("Erro", str(e))
        raise


if __name__ == "__main__":
    main()